"""Parse Bullhorn billable_charges.xlsx export.

Sheet `Page1_1`, header row 1, data starts row 2.

Columns observed:
  A: Employee Type (e.g. 'W2 Hourly')
  B: Invoice Date              <-- billing-week date used as match key
  C: Candidate ID              <-- contractor key
  D: Candidate Name
  E: Client Company Name
  F: Work Location State
  G: Perm or Conv Fee
  H: Total Hours Billed
  I: Reg Bill Rate
  J: Employee Type (dup)
  K: Gross Sales Amount        <-- amount used in matching
  L: Reg Pay Rate
  M: Burdened Wage Amount
  N: Job Title
  O: Gross Margin
  P: Placement ID
  Q: Billing Period End Date
  R: Branch2 (1 = OTS, 2 = CTS — confirm)
"""
from __future__ import annotations
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
import openpyxl
from period_calendar import to_period_anchor

HEADER_ROW = 1
DATA_START_ROW = 2

COL_EMP_TYPE = 1       # A
COL_DATE = 2           # B (Invoice Date)
COL_CANDIDATE_ID = 3   # C
COL_CANDIDATE_NAME = 4 # D
COL_CLIENT = 5         # E
COL_STATE = 6          # F
COL_PERM_FEE = 7       # G
COL_HOURS = 8          # H
COL_BILL_RATE = 9      # I
COL_AMOUNT = 11        # K (Gross Sales Amount)
COL_PAY_RATE = 12      # L
COL_JOB_TITLE = 14     # N
COL_GM = 15            # O
COL_PLACEMENT_ID = 16  # P
COL_PERIOD_END = 17    # Q
COL_BRANCH = 18        # R


@dataclass
class BHRow:
    row_num: int
    invoice_date: date | None
    candidate_id: str
    candidate_name: str
    client: str
    state: str
    hours: float
    bill_rate: float
    gross_sales_amount: float
    pay_rate: float
    job_title: str
    gross_margin: float
    placement_id: str
    period_end: date | None
    branch: str
    employee_type: str

    @property
    def key(self) -> str | None:
        """candidate_id|YYYY-MM-DD using the accounting-period-anchored date."""
        d = to_period_anchor(self.invoice_date)
        if not self.candidate_id or not d:
            return None
        return f"{self.candidate_id}|{d.isoformat()}"


def _to_date(v) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    return None


def _to_float(v) -> float:
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _to_str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def load_bullhorn(path: Path) -> list[BHRow]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    # Pick the first sheet (currently 'Page1_1'; tolerate renames).
    ws = wb[wb.sheetnames[0]]
    rows: list[BHRow] = []
    for row_num, raw in enumerate(
        ws.iter_rows(min_row=DATA_START_ROW, values_only=True), start=DATA_START_ROW
    ):
        cells = list(raw) + [None] * (18 - len(raw)) if len(raw) < 18 else list(raw)

        candidate_id = _to_str(cells[COL_CANDIDATE_ID - 1])
        invoice_date = _to_date(cells[COL_DATE - 1])

        # Skip rows lacking either side of the match key (also catches blank tail rows).
        if not candidate_id or not invoice_date:
            continue

        rows.append(
            BHRow(
                row_num=row_num,
                invoice_date=invoice_date,
                candidate_id=candidate_id,
                candidate_name=_to_str(cells[COL_CANDIDATE_NAME - 1]),
                client=_to_str(cells[COL_CLIENT - 1]),
                state=_to_str(cells[COL_STATE - 1]),
                hours=_to_float(cells[COL_HOURS - 1]),
                bill_rate=_to_float(cells[COL_BILL_RATE - 1]),
                gross_sales_amount=round(_to_float(cells[COL_AMOUNT - 1]), 2),
                pay_rate=_to_float(cells[COL_PAY_RATE - 1]),
                job_title=_to_str(cells[COL_JOB_TITLE - 1]),
                gross_margin=_to_float(cells[COL_GM - 1]),
                placement_id=_to_str(cells[COL_PLACEMENT_ID - 1]),
                period_end=_to_date(cells[COL_PERIOD_END - 1]),
                branch=_to_str(cells[COL_BRANCH - 1]),
                employee_type=_to_str(cells[COL_EMP_TYPE - 1]),
            )
        )

    wb.close()
    return rows


if __name__ == "__main__":
    import sys
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(
        r"C:\Users\pwil\OakTree Software, Inc d b a OakTree Staffing\Management - PBW Agents\OakTree Operations\_shared\bullhorn-data\billable_charges.xlsx"
    )
    rows = load_bullhorn(src)
    print(f"Parsed {len(rows)} Bullhorn rows from {src.name}")
    if rows:
        # Date range
        dates = sorted({r.invoice_date for r in rows})
        print(f"  date range: {dates[0]} ... {dates[-1]}  ({len(dates)} unique dates)")
        # Distinct branches
        branches = sorted({r.branch for r in rows})
        print(f"  branches: {branches}")
        # April 2026 slice
        from datetime import date as D
        apr = [r for r in rows if r.invoice_date and r.invoice_date.year == 2026 and r.invoice_date.month == 4]
        print(f"  April 2026 rows: {len(apr)}")
        print(f"  example: {rows[0]}")
