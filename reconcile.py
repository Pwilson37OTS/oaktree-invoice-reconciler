"""Bullhorn ↔ QBO Revenue Reconciler — orchestrator.

Per-entity (OTS / CTS) end-to-end reconciliation. Reads the two source files,
runs the same match logic as the legacy VBA macro, classifies, and writes:

  output/{entity}_invoice_reconciliation_data.json   — full dataset for Streamlit
  output/audit_{entity}_YYYY-MM.xlsx                 — sign-off-ready snapshot

`--entity all` (default) runs both OTS and CTS back-to-back.
"""
from __future__ import annotations
import argparse
import json
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

from parse_qbo import load_qbo
from parse_bullhorn import load_bullhorn
from match import reconcile, ReconciledRow
from classify import (
    classify, status_counts,
    STATUS_CLEAN, STATUS_PENNY, STATUS_ZERO_NO_ACTION, STATUS_EXPENSE_RECLASS,
)
from period_calendar import to_period_anchor

HERE = Path(__file__).resolve().parent
ROOT = HERE.parents[1]  # OakTree Operations
REPO_DATA = HERE / "data"          # bundled with the repo (cloud deploy)
OUT_DIR = HERE / "output"
OUT_DIR.mkdir(parents=True, exist_ok=True)

# Entities whose billing dates should be remapped through period_calendar
# (Jan 31 / June 1 / etc. cutoffs). CTS bills cleanly on Sundays with no
# period-end pulling, so it stays out of this set.
ENTITIES_WITH_PERIOD_REMAP: set[str] = {"ots"}

# Each entity has an ordered list of candidate paths. The first existing path
# is used. Order: local OneDrive first (live data on Phil's machine), then
# repo-bundled data/ (works on Streamlit Cloud). This means locally you keep
# seeing fresh OneDrive-synced data; in the cloud you see whatever was last
# committed to data/.
ENTITY_CONFIG: dict[str, dict] = {
    "ots": {
        "label": "OTS",
        "qbo_candidates": [
            ROOT / "_shared" / "quickbooks-data" / "audit_rev_ots.xlsx",
            REPO_DATA / "audit_rev_ots.xlsx",
        ],
        "ats_candidates": [
            ROOT / "_shared" / "bullhorn-data" / "billable_charges.xlsx",
            REPO_DATA / "billable_charges.xlsx",
        ],
    },
    "cts": {
        "label": "CTS",
        "qbo_candidates": [
            ROOT / "_shared" / "quickbooks-data" / "audit_rev_cts.xlsx",
            REPO_DATA / "audit_rev_cts.xlsx",
        ],
        "ats_candidates": [
            ROOT / "_shared" / "bullhorn-data" / "billable_charges_CTS.xlsx",
            REPO_DATA / "billable_charges_CTS.xlsx",
        ],
    },
}


def _first_existing(paths: list[Path]) -> Path | None:
    """Return the first path that exists, or None."""
    for p in paths:
        if p.exists():
            return p
    return None


def entity_paths(entity: str) -> tuple[Path | None, Path | None]:
    """Resolve the (qbo, ats) source paths for an entity by checking
    candidates in order."""
    cfg = ENTITY_CONFIG[entity]
    return _first_existing(cfg["qbo_candidates"]), _first_existing(cfg["ats_candidates"])


def _effective_date(row):
    """Return the date used by the match key. Mirrors the row's own .key
    logic — period-anchored only when the row's apply_period_remap flag is
    set (OTS yes, CTS no)."""
    apply = getattr(row, "apply_period_remap", True)
    for attr in ("parsed_date", "txn_date", "invoice_date"):
        d = getattr(row, attr, None)
        if d:
            return to_period_anchor(d) if apply else d
    return None


def _filter_by_month(rows, month: str | None):
    """Restrict `rows` to those whose effective key-date falls in month=YYYY-MM."""
    if not month:
        return rows
    yr, mo = (int(p) for p in month.split("-"))
    out = []
    for r in rows:
        d = _effective_date(r)
        if d and d.year == yr and d.month == mo:
            out.append(r)
    return out


def _row_to_json(r: ReconciledRow) -> dict:
    return {
        "key": r.key,
        "candidate_id": r.candidate_id,
        "billing_date": r.billing_date.isoformat(),
        "contractor": r.contractor_name,
        "client": r.client,
        "placement_id": r.placement_id,
        "job_title": r.job_title,
        "qbo_amount": r.qbo_amount,
        "ats_amount": r.ats_amount,
        "diff": r.diff,
        "status": r.status,
        "qbo_lines": [
            {
                "row": q.row_num,
                "txn_date": q.txn_date.isoformat() if q.txn_date else None,
                "txn_type": q.txn_type,
                "num": q.num,
                "customer": q.customer,
                "description": q.description,
                "account": q.account,
                "split_account": q.split_account,
                "product": q.product,
                "amount": q.amount,
            }
            for q in r.qbo_rows
        ],
        "ats_lines": [
            {
                "row": a.row_num,
                "invoice_date": a.invoice_date.isoformat() if a.invoice_date else None,
                "hours": a.hours,
                "bill_rate": a.bill_rate,
                "amount": a.gross_sales_amount,
                "client": a.client,
                "placement_id": a.placement_id,
                "job_title": a.job_title,
                "branch": a.branch,
                "employee_type": a.employee_type,
            }
            for a in r.ats_rows
        ],
    }


def write_excel_snapshot(rows: list[ReconciledRow], path: Path, period_label: str, entity_label: str) -> None:
    wb = openpyxl.Workbook()
    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DCE6F1")
    disc_fill = PatternFill("solid", fgColor="FFE4E1")

    # ---- Summary tab ----
    summary = wb.active
    summary.title = "Summary"
    summary["A1"] = f"Bullhorn ↔ QBO Revenue Audit — {entity_label}"
    summary["A1"].font = Font(bold=True, size=14)
    summary["A2"] = f"Period: {period_label}"
    summary["A3"] = f"Generated: {datetime.now().isoformat(timespec='seconds')}"

    counts = status_counts(rows)
    summary["A5"] = "Status"
    summary["B5"] = "Count"
    summary["A5"].font = summary["B5"].font = bold
    for i, (status, cnt) in enumerate(sorted(counts.items()), start=6):
        summary.cell(row=i, column=1, value=status)
        summary.cell(row=i, column=2, value=cnt)

    qbo_total = round(sum(r.qbo_amount or 0 for r in rows), 2)
    ats_total = round(sum(r.ats_amount or 0 for r in rows), 2)
    net = round(ats_total - qbo_total, 2)
    base_row = 6 + len(counts) + 1
    summary.cell(row=base_row, column=1, value="QBO total").font = bold
    summary.cell(row=base_row, column=2, value=qbo_total)
    summary.cell(row=base_row + 1, column=1, value="ATS total").font = bold
    summary.cell(row=base_row + 1, column=2, value=ats_total)
    summary.cell(row=base_row + 2, column=1, value="Net variance (ATS - QBO)").font = bold
    summary.cell(row=base_row + 2, column=2, value=net)

    # ---- Audit tab (matches VBA workbook column shape) ----
    audit = wb.create_sheet("Audit")
    headers = ["Contractor Name", "QBO Date", "QBO Amount", "ATS Date", "ATS Amount", "Difference (ATS - QBO)", "Status"]
    for col, h in enumerate(headers, start=1):
        c = audit.cell(row=1, column=col, value=h)
        c.font = bold
        c.fill = header_fill

    for i, r in enumerate(rows, start=2):
        audit.cell(row=i, column=1, value=r.contractor_name)
        audit.cell(row=i, column=2, value=r.billing_date if r.qbo_amount is not None else None)
        audit.cell(row=i, column=3, value=r.qbo_amount)
        audit.cell(row=i, column=4, value=r.billing_date if r.ats_amount is not None else None)
        audit.cell(row=i, column=5, value=r.ats_amount)
        audit.cell(row=i, column=6, value=r.diff)
        audit.cell(row=i, column=7, value=r.status)

    # ---- Discrepancies tab ----
    disc = wb.create_sheet("Discrepancies")
    for col, h in enumerate(headers, start=1):
        c = disc.cell(row=1, column=col, value=h)
        c.font = bold
        c.fill = header_fill

    drow = 2
    NON_ACTIONABLE = (STATUS_CLEAN, STATUS_PENNY, STATUS_ZERO_NO_ACTION, STATUS_EXPENSE_RECLASS)
    for r in rows:
        if r.status in NON_ACTIONABLE:
            continue
        disc.cell(row=drow, column=1, value=r.contractor_name)
        disc.cell(row=drow, column=2, value=r.billing_date if r.qbo_amount is not None else None)
        disc.cell(row=drow, column=3, value=r.qbo_amount)
        disc.cell(row=drow, column=4, value=r.billing_date if r.ats_amount is not None else None)
        disc.cell(row=drow, column=5, value=r.ats_amount)
        disc.cell(row=drow, column=6, value=r.diff)
        disc.cell(row=drow, column=7, value=r.status)
        for c in range(1, 8):
            disc.cell(row=drow, column=c).fill = disc_fill
        drow += 1

    for sheet in (audit, disc, summary):
        for col_idx in range(1, sheet.max_column + 1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 22

    wb.save(path)


def reconcile_entity(entity: str, *, month: str | None = None,
                     qbo_path: Path | None = None, ats_path: Path | None = None,
                     out_json: Path | None = None, out_xlsx: Path | None = None,
                     write_outputs: bool = True, verbose: bool = True) -> dict:
    """Run one entity end-to-end. Returns the payload dict.

    When `write_outputs=True` (CLI default) also writes JSON + Excel to disk.
    When False (Streamlit-Cloud library use) skips disk I/O entirely — caller
    holds the returned dict in memory.
    """
    cfg = ENTITY_CONFIG[entity]
    default_qbo, default_ats = entity_paths(entity)
    qbo_path = qbo_path or default_qbo
    ats_path = ats_path or default_ats
    if not (qbo_path and ats_path):
        raise FileNotFoundError(
            f"No source files found for {entity}. Checked: "
            f"qbo={cfg['qbo_candidates']}, ats={cfg['ats_candidates']}"
        )
    out_json = out_json or (OUT_DIR / f"{entity}_invoice_reconciliation_data.json")
    out_xlsx = out_xlsx or (OUT_DIR / f"audit_{entity}_{month or 'all'}.xlsx")

    def log(msg):
        if verbose:
            print(msg)

    apply_remap = entity in ENTITIES_WITH_PERIOD_REMAP

    log(f"\n=== {cfg['label']} ===  (period_remap={'on' if apply_remap else 'off'})")
    log(f"Loading QBO  : {qbo_path.name}")
    qbo_rows = load_qbo(qbo_path, apply_period_remap=apply_remap)
    log(f"  parsed {len(qbo_rows)} rows")

    log(f"Loading ATS  : {ats_path.name}")
    ats_rows = load_bullhorn(ats_path, apply_period_remap=apply_remap)
    log(f"  parsed {len(ats_rows)} rows")

    if month:
        before = len(ats_rows)
        ats_rows = _filter_by_month(ats_rows, month)
        log(f"  filtered to {month}: {before} -> {len(ats_rows)} ATS rows")
        before = len(qbo_rows)
        qbo_rows = _filter_by_month(qbo_rows, month)
        log(f"  filtered to {month}: {before} -> {len(qbo_rows)} QBO rows")

    log("Reconciling...")
    reconciled = reconcile(qbo_rows, ats_rows)
    classify(reconciled)

    counts = status_counts(reconciled)
    log(f"  reconciled rows: {len(reconciled)}")
    for status, cnt in sorted(counts.items()):
        log(f"    {status}: {cnt}")
    non_actionable = {STATUS_CLEAN, STATUS_PENNY, STATUS_ZERO_NO_ACTION, STATUS_EXPENSE_RECLASS}
    actionable_count = sum(c for s, c in counts.items() if s not in non_actionable)
    log(f"  actionable exceptions: {actionable_count}")

    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "entity": entity,
        "entity_label": cfg["label"],
        "period": month or "all",
        "sources": {"qbo": str(qbo_path), "ats": str(ats_path)},
        "summary": {
            "qbo_total": round(sum(r.qbo_amount or 0 for r in reconciled), 2),
            "ats_total": round(sum(r.ats_amount or 0 for r in reconciled), 2),
            "row_count": len(reconciled),
            "status_counts": counts,
            "actionable_exceptions": actionable_count,
        },
        "rows": [_row_to_json(r) for r in reconciled],
        "_reconciled": reconciled,  # held for in-process snapshot generation; not serialized
    }

    if write_outputs:
        # Strip non-serializable handle before writing
        json_payload = {k: v for k, v in payload.items() if not k.startswith("_")}
        out_json.write_text(json.dumps(json_payload, indent=2, default=str))
        log(f"Wrote JSON   : {out_json}")
        write_excel_snapshot(reconciled, out_xlsx, period_label=month or "all", entity_label=cfg["label"])
        log(f"Wrote Excel  : {out_xlsx}")

    return payload


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--entity", choices=("ots", "cts", "all"), default="all",
                    help="Which entity to reconcile. Default: all (runs OTS then CTS).")
    ap.add_argument("--qbo", type=Path, default=None, help="Override QBO source xlsx (single-entity runs only).")
    ap.add_argument("--ats", type=Path, default=None, help="Override Bullhorn source xlsx (single-entity runs only).")
    ap.add_argument("--month", type=str, default=None, help="YYYY-MM. Filter both sides to this period before joining.")
    args = ap.parse_args()

    entities = ["ots", "cts"] if args.entity == "all" else [args.entity]

    if (args.qbo or args.ats) and args.entity == "all":
        ap.error("--qbo / --ats only valid with a single --entity, not 'all'.")

    for ent in entities:
        reconcile_entity(
            ent,
            month=args.month,
            qbo_path=args.qbo,
            ats_path=args.ats,
        )


if __name__ == "__main__":
    main()
