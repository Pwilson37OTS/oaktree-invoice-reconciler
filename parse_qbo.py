"""Parse a QBO Revenue Audit Report export — handles both layouts.

OakTree publishes two slightly different QBO report templates:

  OTS layout ('Revenue Audit Report_Includes Expenses'):
    A: (account-section header, e.g. "Employee Advance")
    B: Transaction date    C: Transaction type    D: Num    E: Name
    F: Description         G: Full name (account) H: Item split account
    I: Amount              J: Balance

  CTS layout ('Revenue Audit - All Sales/Invoices'):
    A: Transaction ID      B: Transaction date    C: Transaction type
    D: Num                 E: Date                F: Customer
    G: Product/Service     H: Description         I: Distribution account
    J: Amount

Rather than hardcode column positions, the parser reads the header row and
maps logical fields to physical columns by header name. This way the same
parser handles both templates (and any future column reshuffle).

Pre-processing performed (mirrors the VBA macro in the legacy workbook):
1. Walk rows after the header row.
2. From the Description, extract candidate_id (first comma-token) and
   billing-week date (rightmost ` - ` token that IsDate-parses).
3. Backfill blank Transaction-date cells by walking upward (QBO groups
   multi-line invoices under one date in some templates).
4. Sign-flip Employee Advance lines (when present): Credit Memo => -abs,
   all other transaction types => +abs.
"""
from __future__ import annotations
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
import re
import openpyxl
from period_calendar import to_period_anchor


# Logical column -> list of acceptable header strings (lowercased, stripped).
# First match wins. Add aliases as new QBO export templates surface.
COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "date":     ("transaction date",),
    "txn_type": ("transaction type",),
    "num":      ("num",),
    "customer": ("name", "customer"),
    "description": ("description",),
    "account":  ("full name", "account name", "distribution account"),
    "split":    ("item split account",),    # OTS only; missing in CTS template
    "product":  ("product/service",),        # CTS only; drives CPL2 sign rule
    "amount":   ("amount",),
}

# QBO Product/Service codes whose negative amounts should be flipped to
# positive. These are entries booked outside the revenue account that
# nevertheless represent positive revenue for reconciliation purposes.
# Phil-confirmed:
#   CPL2 — CTS line items posted as negatives because they're booked
#          somewhere other than revenue. Treat as positive.
# Extend by adding the product code (uppercase) to this set.
SIGN_FLIP_PRODUCTS: set[str] = {"CPL2"}

# Where to look for the header row. Both current templates use row 5.
DEFAULT_HEADER_ROW = 5


@dataclass
class QBORow:
    row_num: int
    txn_date: date | None
    txn_type: str
    num: str
    customer: str
    description: str
    account: str
    split_account: str
    amount: float
    candidate_id: str | None
    parsed_date: date | None
    product: str = ""
    apply_period_remap: bool = True  # entities like CTS pass False — see reconcile.ENTITIES_WITH_PERIOD_REMAP

    @property
    def key(self) -> str | None:
        """candidate_id|YYYY-MM-DD. Anchors cross-period billings (e.g.
        2026-05-31 → 2026-06-01) when apply_period_remap=True, otherwise
        uses the raw date as-is."""
        d = self.parsed_date or self.txn_date
        if self.apply_period_remap:
            d = to_period_anchor(d)
        if not self.candidate_id or not d:
            return None
        return f"{self.candidate_id}|{d.isoformat()}"


def _to_date(v) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    return None


def _parse_description(desc: str) -> tuple[str | None, date | None]:
    """Return (candidate_id, billing_week_date) extracted from description.

    Description formats observed:
      Standard:  "182449, Anselmo Morales - Regular - 2025-12-14"
      Extended:  "182827, Erika Rivera - Regular - 2025-12-13 - El Paso - TX - 79930"
      Expense:   "58363, Angela Hardin - Expense - 2026-04-03"
      Mileage:   "58363, Angela Hardin - Mileage - 2026-04-03"

    candidate_id = first comma-token (or whole trimmed desc if no comma).
    date         = rightmost ` - ` token that parses as a date.
    """
    if not desc:
        return None, None
    desc = desc.strip()
    if not desc:
        return None, None

    if "," in desc:
        candidate_id = desc.split(",", 1)[0].strip()
    else:
        candidate_id = desc

    billing_date: date | None = None
    parts = desc.split(" - ")
    # Walk right-to-left like the VBA does.
    for part in reversed(parts):
        d = _to_date(part.strip())
        if d:
            billing_date = d
            break

    return (candidate_id or None), billing_date


def _sign_correct(account: str, txn_type: str, product: str, amount: float) -> float:
    """Normalize amount signs so per-key sums net correctly:

    1. Product/Service in SIGN_FLIP_PRODUCTS (e.g. CPL2): always +abs.
       These are booked outside revenue and post as negatives but should
       be treated as positive revenue for reconciliation.
    2. Employee Advance lines: Credit Memo => -abs, all else => +abs.
       Mirrors the VBA macro's normalization.
    Product rule takes precedence when both apply.
    """
    if product and product.strip().upper() in SIGN_FLIP_PRODUCTS:
        return abs(amount)
    if (account or "").strip() == "Employee Advance":
        if (txn_type or "").strip() == "Credit Memo":
            return -abs(amount)
        return abs(amount)
    return amount


def _resolve_columns(ws, header_row: int) -> dict[str, int]:
    """Read the header row and map logical field name -> 1-indexed column."""
    headers_row = next(iter(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)), ())
    by_header: dict[str, int] = {}
    for i, h in enumerate(headers_row, start=1):
        if h is None:
            continue
        by_header[str(h).strip().lower()] = i

    resolved: dict[str, int] = {}
    for logical, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in by_header:
                resolved[logical] = by_header[alias]
                break
    return resolved


def _cell(cells, col_idx: int | None):
    """Safe accessor — returns None if column is missing from this template."""
    if col_idx is None:
        return None
    if col_idx - 1 >= len(cells):
        return None
    return cells[col_idx - 1]


def load_qbo(path: Path, header_row: int = DEFAULT_HEADER_ROW,
             apply_period_remap: bool = True) -> list[QBORow]:
    """Load a QBO Revenue Audit Report. Auto-detects column layout from header.

    apply_period_remap: if False, raw dates are kept (used for CTS, which
    bills weekly on Sundays with no period-end pulling).
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    cols = _resolve_columns(ws, header_row)
    # `account` is required: it drives Employee Advance sign-correction and the
    # P&L / account-breakdown tiles. A silent blank account (e.g. from a header
    # rename like "Full name" -> "Account Name") corrupts the reconciliation
    # math without any error, so we fail loudly instead. If QBO renames this
    # column again, add the new header to COLUMN_ALIASES["account"] above.
    required = ("date", "description", "amount", "account")
    missing = [k for k in required if k not in cols]
    if missing:
        header_vals = [str(c).strip() if c else "" for c in
                       next(iter(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)), ())]
        wb.close()
        raise ValueError(
            f"QBO file {path.name!r} header row {header_row} is missing required "
            f"column(s) {missing}. Actual headers found: {header_vals}. "
            f"If a column was renamed, add its new name to COLUMN_ALIASES in parse_qbo.py."
        )

    rows: list[QBORow] = []
    last_known_date: date | None = None

    for row_num, raw in enumerate(
        ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1
    ):
        cells = list(raw)

        desc_raw = _cell(cells, cols.get("description")) or ""
        desc = desc_raw.strip() if isinstance(desc_raw, str) else str(desc_raw).strip()

        # Skip non-data rows (section headers, footers, blanks).
        if not desc:
            d = _to_date(_cell(cells, cols.get("date")))
            if d:
                last_known_date = d
            continue

        candidate_id, parsed_date = _parse_description(desc)

        # txn_date = the actual Transaction Date column (P&L / booking date).
        # Backfills from the previous data row when blank (QBO groups multi-
        # line invoices under one date). Used for the P&L-month view.
        # parsed_date (from Description) is kept separately and used for the
        # match key — that's the service-week date and ties to Bullhorn.
        txn_date = _to_date(_cell(cells, cols.get("date"))) or last_known_date
        if txn_date:
            last_known_date = txn_date

        txn_type = str(_cell(cells, cols.get("txn_type")) or "").strip()
        num = str(_cell(cells, cols.get("num")) or "").strip()
        customer = str(_cell(cells, cols.get("customer")) or "").strip()
        account = str(_cell(cells, cols.get("account")) or "").strip()
        split_account = str(_cell(cells, cols.get("split")) or "").strip()
        product = str(_cell(cells, cols.get("product")) or "").strip()

        amount_raw = _cell(cells, cols.get("amount"))
        try:
            amount = float(amount_raw) if amount_raw not in (None, "") else 0.0
        except (TypeError, ValueError):
            # Footer rows can land here (timestamps). Skip.
            continue

        amount = _sign_correct(account, txn_type, product, amount)

        rows.append(
            QBORow(
                row_num=row_num,
                txn_date=txn_date,
                txn_type=txn_type,
                num=num,
                customer=customer,
                description=desc,
                account=account,
                split_account=split_account,
                amount=amount,
                candidate_id=candidate_id,
                parsed_date=parsed_date,
                product=product,
                apply_period_remap=apply_period_remap,
            )
        )

    wb.close()
    return rows


if __name__ == "__main__":
    import sys
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(
        r"C:\Users\pwil\OakTree Software, Inc d b a OakTree Staffing\Management - PBW Agents\OakTree Operations\_shared\quickbooks-data\audit_rev_ots.xlsx"
    )
    rows = load_qbo(src)
    print(f"Parsed {len(rows)} QBO rows from {src.name}")
    keyed = sum(1 for r in rows if r.key)
    print(f"  with valid match key: {keyed}")
    print(f"  example: {rows[0] if rows else None}")
